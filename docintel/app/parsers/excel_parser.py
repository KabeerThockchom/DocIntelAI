import os
import time
import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, BinaryIO, Tuple
from io import BytesIO

from app.parsers.base_parser import BaseDocumentParser
from app.chunking.models import ProcessedDocument
from app.chunking.chunker import DocumentChunker
from app.utils.logging import log_step, Timer


class ExcelParser(BaseDocumentParser):
    """Parser for Excel and CSV documents."""
    
    def __init__(self, chunker: Optional[DocumentChunker] = None):
        """
        Initialize Excel parser.
        
        Args:
            chunker: Document chunker instance (optional)
        """
        super().__init__()
        self.chunker = chunker or DocumentChunker()
    
    def parse(
        self, 
        file_path: str, 
        filename: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ProcessedDocument:
        """
        Parse an Excel or CSV file.
        
        Args:
            file_path: Path to the Excel/CSV file
            filename: Original filename (uses basename of file_path if not provided)
            metadata: Additional metadata
            
        Returns:
            ProcessedDocument object with extracted content and metadata
        """
        with Timer("Excel Parsing"):
            # Get filename if not provided
            if not filename:
                filename = os.path.basename(file_path)
                
            log_step("Excel Parsing", f"Parsing Excel/CSV file: {filename}")
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Prepare metadata
            doc_metadata = self.prepare_metadata(filename, file_size, metadata)
            
            # Start timer for processing
            start_time = time.time()
            
            try:
                # Extract text based on file extension
                file_ext = os.path.splitext(filename)[1].lower().lstrip(".")
                
                if file_ext == "csv":
                    text_by_sheet = self._extract_text_from_csv(file_path)
                else:  # xlsx, xls
                    text_by_sheet = self._extract_text_from_excel(file_path)
                
                # No text extracted
                if not text_by_sheet:
                    log_step("Excel Parsing", "No text extracted from Excel/CSV", level="warning")
                    return ProcessedDocument(
                        document_id=self.document_id,
                        filename=filename,
                        file_type=file_ext,
                        file_size=file_size,
                        total_pages=0,
                        total_chunks=0,
                        chunks=[],
                        processing_time=time.time() - start_time,
                        is_complex=False
                    )
                
                # Process each sheet
                all_chunks = []
                
                for sheet_name, sheet_text in text_by_sheet.items():
                    # Skip empty sheets
                    if not sheet_text:
                        continue
                    
                    # Add sheet info to metadata
                    sheet_metadata = doc_metadata.copy()
                    sheet_metadata["sheet_name"] = sheet_name
                    
                    # Chunk the sheet
                    sheet_chunks = self.chunker.chunk_document(
                        sheet_text,
                        sheet_metadata,
                        use_headings=False,  # Excel sheets don't typically have headings in the text
                        is_ocr=False
                    )
                    
                    all_chunks.extend(sheet_chunks)
                
                # Create processed document
                processed_doc = ProcessedDocument(
                    document_id=self.document_id,
                    filename=filename,
                    file_type=file_ext,
                    file_size=file_size,
                    total_pages=len(text_by_sheet),
                    total_chunks=len(all_chunks),
                    chunks=all_chunks,
                    processing_time=time.time() - start_time,
                    is_complex=False
                )
                
                log_step("Excel Parsing", f"Completed parsing Excel/CSV with {len(text_by_sheet)} sheets and {len(all_chunks)} chunks")
                return processed_doc
                
            except Exception as e:
                log_step("Excel Parsing", f"Error parsing Excel/CSV: {str(e)}", level="error")
                raise
    
    def parse_stream(
        self, 
        file_stream: BinaryIO,
        filename: str,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ProcessedDocument:
        """
        Parse an Excel/CSV from a file stream.
        
        Args:
            file_stream: File-like object containing the Excel/CSV
            filename: Original filename
            metadata: Additional metadata
            
        Returns:
            ProcessedDocument object with extracted content and metadata
        """
        with Timer("Excel Stream Parsing"):
            log_step("Excel Parsing", f"Parsing Excel/CSV from stream: {filename}")
            
            # Get file size
            file_stream.seek(0, os.SEEK_END)
            file_size = file_stream.tell()
            file_stream.seek(0)
            
            # Prepare metadata
            doc_metadata = self.prepare_metadata(filename, file_size, metadata)
            
            # Start timer for processing
            start_time = time.time()
            
            try:
                # Extract text based on file extension
                file_ext = os.path.splitext(filename)[1].lower().lstrip(".")
                
                # Create in-memory file
                file_data = file_stream.read()
                memory_stream = BytesIO(file_data)
                
                if file_ext == "csv":
                    text_by_sheet = self._extract_text_from_csv_stream(memory_stream)
                else:  # xlsx, xls
                    text_by_sheet = self._extract_text_from_excel_stream(memory_stream)
                
                # No text extracted
                if not text_by_sheet:
                    log_step("Excel Parsing", "No text extracted from Excel/CSV", level="warning")
                    return ProcessedDocument(
                        document_id=self.document_id,
                        filename=filename,
                        file_type=file_ext,
                        file_size=file_size,
                        total_pages=0,
                        total_chunks=0,
                        chunks=[],
                        processing_time=time.time() - start_time,
                        is_complex=False
                    )
                
                # Process each sheet
                all_chunks = []
                
                for sheet_name, sheet_text in text_by_sheet.items():
                    # Skip empty sheets
                    if not sheet_text:
                        continue
                    
                    # Add sheet info to metadata
                    sheet_metadata = doc_metadata.copy()
                    sheet_metadata["sheet_name"] = sheet_name
                    
                    # Chunk the sheet
                    sheet_chunks = self.chunker.chunk_document(
                        sheet_text,
                        sheet_metadata,
                        use_headings=False,  # Excel sheets don't typically have headings in the text
                        is_ocr=False
                    )
                    
                    all_chunks.extend(sheet_chunks)
                
                # Create processed document
                processed_doc = ProcessedDocument(
                    document_id=self.document_id,
                    filename=filename,
                    file_type=file_ext,
                    file_size=file_size,
                    total_pages=len(text_by_sheet),
                    total_chunks=len(all_chunks),
                    chunks=all_chunks,
                    processing_time=time.time() - start_time,
                    is_complex=False
                )
                
                log_step("Excel Parsing", f"Completed parsing Excel/CSV from stream with {len(text_by_sheet)} sheets and {len(all_chunks)} chunks")
                return processed_doc
                
            except Exception as e:
                log_step("Excel Parsing", f"Error parsing Excel/CSV stream: {str(e)}", level="error")
                raise
    
    def _extract_text_from_csv(self, file_path: str) -> Dict[str, str]:
        """
        Extract text from CSV using pandas.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Dictionary mapping sheet name to extracted text
        """
        try:
            # Read CSV with pandas
            df = pd.read_csv(file_path, encoding='utf-8')
            
            # Convert DataFrame to string representation
            text = df.to_string(index=False)
            
            # Include basic stats
            stats = f"""
            Total Rows: {len(df)}
            Total Columns: {len(df.columns)}
            Column Names: {', '.join(df.columns.tolist())}
            """
            
            text = stats + "\n\n" + text
            
            return {"Sheet1": text}
            
        except Exception as e:
            log_step("Excel Parsing", f"Error extracting text from CSV: {str(e)}", level="warning")
            
            # Try with different encoding
            try:
                df = pd.read_csv(file_path, encoding='latin1')
                text = df.to_string(index=False)
                stats = f"""
                Total Rows: {len(df)}
                Total Columns: {len(df.columns)}
                Column Names: {', '.join(df.columns.tolist())}
                """
                text = stats + "\n\n" + text
                return {"Sheet1": text}
            except Exception as e2:
                log_step("Excel Parsing", f"Error extracting text from CSV with latin1 encoding: {str(e2)}", level="error")
                return {}
    
    def _extract_text_from_excel(self, file_path: str) -> Dict[str, str]:
        """
        Extract text from Excel using openpyxl and pandas.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary mapping sheet names to extracted text
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract text from each sheet
            text_by_sheet = {}
            
            for sheet_name in workbook.sheetnames:
                # Read sheet with pandas
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Convert DataFrame to string representation
                sheet_text = df.to_string(index=False)
                
                # Include basic stats
                stats = f"""
                Sheet: {sheet_name}
                Total Rows: {len(df)}
                Total Columns: {len(df.columns)}
                Column Names: {', '.join(df.columns.tolist())}
                """
                
                sheet_text = stats + "\n\n" + sheet_text
                
                text_by_sheet[sheet_name] = sheet_text
            
            return text_by_sheet
            
        except Exception as e:
            log_step("Excel Parsing", f"Error extracting text from Excel: {str(e)}", level="warning")
            
            # Fallback to pandas-only approach
            try:
                # Get all sheet names
                excel = pd.ExcelFile(file_path)
                sheet_names = excel.sheet_names
                
                text_by_sheet = {}
                
                for sheet_name in sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    sheet_text = df.to_string(index=False)
                    
                    stats = f"""
                    Sheet: {sheet_name}
                    Total Rows: {len(df)}
                    Total Columns: {len(df.columns)}
                    Column Names: {', '.join(df.columns.tolist())}
                    """
                    
                    sheet_text = stats + "\n\n" + sheet_text
                    
                    text_by_sheet[sheet_name] = sheet_text
                
                return text_by_sheet
                
            except Exception as e2:
                log_step("Excel Parsing", f"Error in fallback Excel parsing: {str(e2)}", level="error")
                return {}
    
    def _extract_text_from_csv_stream(self, memory_stream: BytesIO) -> Dict[str, str]:
        """
        Extract text from CSV stream using pandas.
        
        Args:
            memory_stream: BytesIO object containing the CSV
            
        Returns:
            Dictionary mapping sheet name to extracted text
        """
        try:
            # Reset stream position
            memory_stream.seek(0)
            
            # Read CSV with pandas
            df = pd.read_csv(memory_stream, encoding='utf-8')
            
            # Convert DataFrame to string representation
            text = df.to_string(index=False)
            
            # Include basic stats
            stats = f"""
            Total Rows: {len(df)}
            Total Columns: {len(df.columns)}
            Column Names: {', '.join(df.columns.tolist())}
            """
            
            text = stats + "\n\n" + text
            
            return {"Sheet1": text}
            
        except Exception as e:
            log_step("Excel Parsing", f"Error extracting text from CSV stream: {str(e)}", level="warning")
            
            # Try with different encoding
            try:
                memory_stream.seek(0)
                df = pd.read_csv(memory_stream, encoding='latin1')
                text = df.to_string(index=False)
                stats = f"""
                Total Rows: {len(df)}
                Total Columns: {len(df.columns)}
                Column Names: {', '.join(df.columns.tolist())}
                """
                text = stats + "\n\n" + text
                return {"Sheet1": text}
            except Exception as e2:
                log_step("Excel Parsing", f"Error extracting text from CSV stream with latin1 encoding: {str(e2)}", level="error")
                return {}
    
    def _extract_text_from_excel_stream(self, memory_stream: BytesIO) -> Dict[str, str]:
        """
        Extract text from Excel stream using openpyxl and pandas.
        
        Args:
            memory_stream: BytesIO object containing the Excel file
            
        Returns:
            Dictionary mapping sheet names to extracted text
        """
        try:
            # Reset stream position
            memory_stream.seek(0)
            
            # Load workbook
            workbook = openpyxl.load_workbook(memory_stream, data_only=True)
            
            # Extract text from each sheet
            text_by_sheet = {}
            
            for sheet_name in workbook.sheetnames:
                # Reset stream position for each sheet
                memory_stream.seek(0)
                
                # Read sheet with pandas
                df = pd.read_excel(memory_stream, sheet_name=sheet_name)
                
                # Convert DataFrame to string representation
                sheet_text = df.to_string(index=False)
                
                # Include basic stats
                stats = f"""
                Sheet: {sheet_name}
                Total Rows: {len(df)}
                Total Columns: {len(df.columns)}
                Column Names: {', '.join(df.columns.tolist())}
                """
                
                sheet_text = stats + "\n\n" + sheet_text
                
                text_by_sheet[sheet_name] = sheet_text
            
            return text_by_sheet
            
        except Exception as e:
            log_step("Excel Parsing", f"Error extracting text from Excel stream: {str(e)}", level="warning")
            
            # Fallback to pandas-only approach
            try:
                # Reset stream position
                memory_stream.seek(0)
                
                # Get all sheet names
                excel = pd.ExcelFile(memory_stream)
                sheet_names = excel.sheet_names
                
                text_by_sheet = {}
                
                for sheet_name in sheet_names:
                    # Reset stream position for each sheet
                    memory_stream.seek(0)
                    
                    df = pd.read_excel(memory_stream, sheet_name=sheet_name)
                    sheet_text = df.to_string(index=False)
                    
                    stats = f"""
                    Sheet: {sheet_name}
                    Total Rows: {len(df)}
                    Total Columns: {len(df.columns)}
                    Column Names: {', '.join(df.columns.tolist())}
                    """
                    
                    sheet_text = stats + "\n\n" + sheet_text
                    
                    text_by_sheet[sheet_name] = sheet_text
                
                return text_by_sheet
                
            except Exception as e2:
                log_step("Excel Parsing", f"Error in fallback Excel stream parsing: {str(e2)}", level="error")
                return {}